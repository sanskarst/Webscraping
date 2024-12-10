import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active


ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

ws['A1'] = 'Invoice'

#ws['A1'].font = Font(name='Times New Roman', size=24, bold = True)
#or you could create an object:
fontObj = Font(name='Times New Roman', size=24, bold = True,italic=True)

ws['A1'].font = fontObj

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = fontObj

ws.merge_cells('A1:B1')

ws['B8'] = '=SUM(B2:B4)'

ws.column_dimensions['A'].width = 25


write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']
row = 1
for currentrow in read_ws.iter_rows(min_row=1, max_row=read_ws.max_row,max_col=read_ws.max_column):
    name = currentrow[0].value
    cost = currentrow[1].value
    amtsold = currentrow[2].value
    total = currentrow[3].value
    write_sheet.cell(row=row,column=1,value=name)
    write_sheet.cell(row=row,column=2,value=cost)
    write_sheet.cell(row=row,column=3,value=amtsold)
    write_sheet.cell(row=row,column=4,value=total)
    row+=1
write_sheet['B44'] = 'Average'
write_sheet['B44'].font = fontObj
write_sheet['C44'] = '=AVERAGE(C2:C41)'
write_sheet['D44'] = '=AVERAGE(D2:D41)'
write_sheet['B43'] = 'Grand Total'
write_sheet['B43'].font = fontObj
write_sheet['C43'] = '=SUM(C2:C41)'
write_sheet['D43'] = '=SUM(D2:D41)'
write_sheet.column_dimensions['B'].width = 23
write_sheet.column_dimensions['D'].width = 18

#summary_row = row + 1
#write_sheet['B' +str(summary_row)] = 'Total'
for cell in write_sheet['C:C']:
    cell.number_format = '#,##0'

for cell in write_sheet['D:D']:
    cell.number_format = u'"$ "#,##0.00'


wb.save('PythontoExcel.xlsx')
