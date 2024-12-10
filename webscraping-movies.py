
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

movie_rows = soup.findAll('tr')

wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'
ws['A1'] = 'Movie Title'
ws['B1'] = 'Gross Returns'
ws['C1'] = 'No. of Theaters'
ws['D1'] = 'Average/Theater'

for row in range(1,6):
    td = movie_rows[row].findAll('td')
    title = td[1].text
    gross = int(td[5].text.replace('$','').replace(",",""))
    theater = int(td[6].text.replace(",",""))
    avg = round(gross/theater,2)

    ws['A' + str(row+1)].value = title
    ws['B' + str(row+1)].value = gross
    ws['C' + str(row+1)].value = theater
    ws['D' + str(row+1)].value = avg
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

for cell in ws['B:B']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['D:D']:
    cell.number_format = '#,##0'

wb.save('BoxOfficeReport.xlsx')
##
##
##
##

